from flask import Flask, render_template, request, jsonify, send_file
import sqlite3
import pandas as pd
import tempfile
app=Flask(__name__)
DB='database.db'
def c():
    x=sqlite3.connect(DB)
    x.row_factory=sqlite3.Row
    return x
@app.route('/')
def i():
    return render_template('index.html')
@app.route('/buscar')
def b():
    doc=request.args.get('doc','')
    x=c()
    rows=[dict(r) for r in x.execute('SELECT rowid rid,* FROM base WHERE [Doc. Identidad] = ?', (doc,)).fetchall()]
    pisos=[r[0] for r in x.execute('SELECT DISTINCT [Piso] FROM nomenclatura ORDER BY [Piso]').fetchall()]
    x.close()
    return jsonify({'rows':rows,'pisos':pisos})
@app.route('/ubicaciones')
def u():
    piso=request.args.get('piso','')
    x=c()
    data=[r[0] for r in x.execute('SELECT DISTINCT [Ubicación Detallada] FROM nomenclatura WHERE [Piso] = ? ORDER BY [Ubicación Detallada]',(piso,)).fetchall()]
    x.close()
    return jsonify(data)
@app.route('/guardar', methods=['POST'])
def g():
    data=request.get_json()
    ups=data.get('updates',[])
    x=c()
    for u in ups:
        x.execute('UPDATE base SET [PISO] = ?, [UBICACIÓN DETALLADA] = ? WHERE rowid = ?',(u['piso'],u['ubicacion'],u['rid']))
    x.commit();x.close()
    return jsonify({'msg':'Datos guardados correctamente'})

# DESCARGAR EXCEL
@app.route('/excel')
def excel():

    x = c()

    df = pd.read_sql_query(
        '''
        SELECT
            [Doc. Identidad],
            [Etiqueta],
            [Descr. Detallada],
            [Departamento],
            [Proyecto],
            [Responsable],
            [Piso],
            [UBICACIÓN DETALLADA]
        FROM base
        ''',
        x
    )

    x.close()

    temp = tempfile.NamedTemporaryFile(
        delete=False,
        suffix='.xlsx'
    )

    with pd.ExcelWriter(
        temp.name,
        engine='openpyxl'
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name='REPORTE'
        )

        wb = writer.book

        ws = writer.sheets['REPORTE']

        from openpyxl.styles import Font, PatternFill

        # ESTILO ENCABEZADO
        fill = PatternFill(
            start_color='D9D9D9',
            end_color='D9D9D9',
            fill_type='solid'
        )

        font = Font(
            bold=True
        )

        # APLICAR ESTILO
        for cell in ws[1]:

            cell.fill = fill

            cell.font = font

        # AJUSTAR ANCHO COLUMNAS
        for col in ws.columns:

            max_length = 0

            column = col[0].column_letter

            for cell in col:

                try:

                    if len(str(cell.value)) > max_length:

                        max_length = len(str(cell.value))

                except:
                    pass

            adjusted_width = max_length + 5

            ws.column_dimensions[column].width = adjusted_width

    return send_file(
        temp.name,
        as_attachment=True,
        download_name='REPORTE.xlsx'
    )

# VALIDAR
@app.route('/validar')
def validar():

    x = c()

    rows = x.execute(
        'SELECT * FROM base'
    ).fetchall()

    x.close()

    html = '''

    <html>

    <head>

        <title>Reporte</title>

        <style>

        body{
            font-family:Arial;
            margin:20px;
        }

        table{
            border-collapse:collapse;
            width:100%;
        }

        th,td{
            border:1px solid #ccc;
            padding:8px;
            font-size:12px;
        }

        th{
            background:#f0f0f0;
            position:sticky;
            top:0;
        }

        .btn{
            background:#1976d2;
            color:white;
            border:none;
            padding:10px 15px;
            border-radius:5px;
            cursor:pointer;
            margin-bottom:20px;
        }

        </style>

    </head>

    <body>

    <h2>Validación Datos almacenados</h2>

    <a href="/excel">

        <button class="btn">

            Descargar Excel

        </button>

    </a>

    <table>

    '''

    if rows:

        cols = rows[0].keys()

        html += '<tr>'

        for c1 in cols:

            html += f'<th>{c1}</th>'

        html += '</tr>'

        for r in rows:

            html += '<tr>'

            for c1 in cols:

                html += f'<td>{r[c1]}</td>'

            html += '</tr>'

    html += '''

    </table>

    </body>

    </html>

    '''

    return html
